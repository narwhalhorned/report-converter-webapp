import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.filters import AutoFilter, CustomFilter
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
import base64
from datetime import datetime, timedelta

st.set_page_config(page_title="8 Weeks Report Converter", 
                   page_icon=":pencil", 
                   layout="wide"
)

#Front end
def local_css(file_name):
    with open(file_name) as f:
        st.markdown(f"<style>{ f.read()}</style>", unsafe_allow_html=True)

#Front end CSS file
local_css("style.css")

st.title('8 Weeks Report Converter')
st.subheader('Drop raw data below')

#First file uploader
uploaded_file = st.file_uploader('Choose a XLSX file', type='xlsx')
if uploaded_file:
    st.markdown('---')
    df = pd.read_excel(uploaded_file, engine='openpyxl', header=1)
    
    #Assign as title, use row 2 and below as dataframe
    title = df.iloc[0, 0]
    df = df[1:]
    
    #Change data type
    df['Start Date'] = df['Start Date'].dt.strftime('%d/%m/%Y')
    df['End Date'] = df['End Date'].dt.strftime('%d/%m/%Y')
    df['Worker End Date'] = df['Worker End Date'].dt.strftime('%d/%m/%Y')

    #Column width formatting
    df_styled = df.style.set_properties(subset=["Resourcing Specialist"], **{'width': '150px'})
    #Bold header formatting
    df_styled = df_styled.set_table_styles([dict(selector='th', props=[('font-weight', 'bold')])])

    #Create new blank worksheets
    writer = pd.ExcelWriter('output.xlsx', engine='openpyxl')
    writer.book = openpyxl.Workbook()

    df_styled.to_excel(writer, sheet_name='Report', index=False)
    
    df_confirmed = pd.DataFrame()
    df_pending_approval = pd.DataFrame()
    df_accepted = pd.DataFrame()
    df_rejected = pd.DataFrame()
    df_penreview = pd.DataFrame()
    df_declined= pd.DataFrame()
    
    
    #Transfer dataframe to excel
    df_confirmed.to_excel(writer, sheet_name='Confirmed', index=False)
    df_pending_approval.to_excel(writer, sheet_name='Pending Approval', index=False)
    df_accepted.to_excel(writer, sheet_name='Created', index=False)
    df_declined.to_excel(writer, sheet_name='Declined', index=False)
    df_penreview.to_excel(writer, sheet_name='Pending Review', index=False)
    df_rejected.to_excel(writer, sheet_name='Rejected', index=False)
    
    #Delete worksheet "Sheet"
    del writer.book['Sheet']
    
    #Auto-adjust column width
    for column in df:
        column_width = max(df[column].astype(str).map(len).max(), len(column))
        col_idx = df.columns.get_loc(column)
        col_letter = openpyxl.utils.get_column_letter(col_idx+1)
        writer.sheets['Report'].column_dimensions[col_letter].width = column_width
        if col_idx >= 26:  # columns beyond Z
            col_letter2 = openpyxl.utils.get_column_letter(col_idx-25) + 'A'
            writer.sheets['Report'].column_dimensions[col_letter2].width = column_width
            if col_idx >= 27:  # columns beyond AA
                col_letter3 = openpyxl.utils.get_column_letter(col_idx-26) + 'B'
                writer.sheets['Report'].column_dimensions[col_letter3].width = column_width
                
    #Auto-adjust row height
    for row in writer.sheets['Report'].rows:
        row_height = 12
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
            if cell.value is not None:
                cell.font = openpyxl.styles.Font(size=9, name='Arial Unicode MS')
            if len(str(cell.value)) > 100:
                row_height = 24
        row_dimensions = writer.sheets['Report'].row_dimensions[row[0].row]
        row_dimensions.height = row_height
    
    #Set column width for columns A, C, H, U, AA and J
    for col in ['A', 'C', 'H', 'U', 'AA']:
        writer.sheets['Report'].column_dimensions[col].width = 20

    for col in ['J']:
        writer.sheets['Report'].column_dimensions[col].width = 72
        
    #Define workbook
    wb = writer.book
    
    #Define colour
    yellow = "00FFFF00"
    
    #Select the first sheet
    first_sheet = wb.worksheets[0]

    #Get the values of the first row(header) of the first sheet
    first_row_values = [cell.value for cell in first_sheet[1]]
    
    #Iterate the sheets
    for sheet in wb.worksheets[1:]:
        #Replace the values of the first row with those from the first sheet
        for index, value in enumerate(first_row_values):
            sheet.cell(row=1, column=index+1).value = value

    #Formatting the headers for every sheet
    for sheet in wb.worksheets:
        for row in sheet:
            for cell in row:
                cell.font = Font(size=9, name='Arial Unicode MS')
                cell.alignment = cell.alignment.copy(wrapText=False)
                cell.border = Border()

            #Set the font of cells in the first row to bold
            for cell in sheet['1']:
                cell.font = Font(size=9, name='Arial Unicode MS', bold=True)
                

            #Set the fill color of cell Z2 to yellow
            sheet['Z1'].fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")
    
    ##Pasting format from first worksheet to other worksheets
    #Copy the column width and row height from the first sheet to the rest of the sheets
    for column in first_sheet.columns:
        column_letter = column[0].column_letter
        for sheet in wb.worksheets[1:]:
            sheet.column_dimensions[column_letter].width = first_sheet.column_dimensions[column_letter].width

    for row in first_sheet.rows:
        row_index = row[0].row
        for sheet in wb.worksheets[1:]:
            sheet.row_dimensions[row_index].height = first_sheet.row_dimensions[row_index].height

    #Iterate over the rest of the sheets
    for sheet in wb.worksheets[1:]:
        # Replace the values of the first row with those from the first sheet
        for index, value in enumerate(first_row_values):
            sheet.cell(row=1, column=index+1).value = value
            
    ws = wb["Report"]

    #Select the worksheet named "Pending Approval"
    tab_names = ["Confirmed","Pending Approval", "Created", "Declined", "Pending Review", "Rejected"]
    #ws_pending = ["Pending Approval"]
    
    #Define the column headers
    headers = [cell.value for cell in ws[1]]

    #Define font style
    font_style = Font(size=9, name="Arial Unicode MS")
    #tab_names = ["Pending Approval", "Created", "Declined", "Pending Review", "Rejected"]

    #Loop through each tab
    for tab_name in tab_names:
        #Check if there are declined rows
        has_filtered_rows = any(tab_name in str(row[21]) for row in ws.iter_rows(min_row=2, values_only=True))
        #Select the current worksheet
        wd = wb[tab_name]

        #Define the column headers
        headers = [cell.value for cell in ws[1]]

        #Define font style
        font_style = Font(size=9, name="Arial Unicode MS")

        #Check if there are filtered rows
        has_filtered_rows = any(tab_name in str(row[21]) for row in ws.iter_rows(min_row=2, values_only=True))

        #Create the worksheet only if there are filtered rows
        if has_filtered_rows:
            ws_filtered = wb.get_sheet_by_name(tab_name)
            if ws_filtered is None:
                ws_filtered = wb.create_sheet(tab_name)

            #Filter rows where "Work Order Status" column matches the tab_name
            filtered_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if tab_name in str(row[21]):
                    filtered_rows.append(row)

            #Write the column headers to the filtered worksheet with font style
            for col_num, header in enumerate(headers, start=1):
                cell = ws_filtered.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(size=9, name='Arial Unicode MS', bold=True)

            #Write the filtered rows to the filtered worksheet with font style
            for row_num, row in enumerate(filtered_rows, start=2):
                for col_num, cell_value in enumerate(row, start=1):
                    cell = ws_filtered.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.font = font_style
                    
            today = datetime.today().date()
            next_sundays = []

        #Calculate the next two Sundays
            current_day = today
            while len(next_sundays) < 2:
                current_day += timedelta(days=1)
                if current_day.weekday() == 6:  # Sunday is represented by 6
                    next_sundays.append(current_day)

                #Loop through the rows of column M
            for cell in ws['M']:
                # Skip the header row
                if cell.row == 1:
                    continue

                #Check if the cell value is not empty
                if not cell.value:
                    continue

                #Convert the cell value to a datetime object
                date_str = str(cell.value)
                date_obj = datetime.strptime(date_str, '%d/%m/%Y').date()

                #Check if the date falls within the specified range
                if date_obj < today or date_obj <= next_sundays[1]:
                    # Set value to "Yes" in column Z
                    ws.cell(row=cell.row, column=26).value = "Yes"
                    # Middle align the cell in column Z
                    ws.cell(row=cell.row, column=26).alignment = Alignment(horizontal='center')
                else:
                    # Set value to blank in column Z
                    ws.cell(row=cell.row, column=26).value = ""
                     
    rows_to_delete = []

    for cell in ws['V']:
        #Skip the header row
        if cell.row == 1:
            continue
        
        if ws.cell(row=cell.row, column=22).value != "Confirmed":
            rows_to_delete.append(cell.row)

    #Reverse the list of row numbers to delete to avoid shifting issues
    rows_to_delete.reverse()

    for row in rows_to_delete:
        ws.delete_rows(row)
        
    for tab_name in tab_names:
        #Check if there are declined rows
        has_filtered_rows = any(tab_name in str(row[21]) for row in ws.iter_rows(min_row=2, values_only=True))
        #Select the current worksheet
        wd = wb[tab_name]
        
        #Convert the worksheet data into a list of lists
        data = list(wd.iter_rows(values_only=True))

        #Sort the data by column 26 (Z) in descending order
        sorted_data = sorted(data[1:], key=lambda x: x[25], reverse=True)

        #Clear the existing data in the worksheet
        wd.delete_rows(2, wd.max_row)

        #Write the sorted data back to the worksheet
        for row in sorted_data:
            wd.append(row)
            
        #Middle align the "Yes" values in column 26
        for row in wd.iter_rows(min_row=2, max_col=26, max_row=wd.max_row):
            cell = row[25]  # Column 26 (Z)
            if cell.value == 'Yes':
                cell.alignment = Alignment(horizontal='center')

            for cell in row:
                cell.font = Font(size=9, name='Arial Unicode MS')
                
        #Check if there's no data in row 2
        if all(cell.value is None for cell in wd[2]):
            wb.remove(wd)  # Delete the worksheet tab
    
    wb.remove(wb['Confirmed'])
    wb = wd.parent
    
st.subheader('Drop previous data below')

vlookup_file = st.file_uploader('Choose a XLSX file', type='xlsx', key='file_uploader')
if vlookup_file:
    st.markdown('---')

    #Open the vlookup_file workbook
    wvl = openpyxl.load_workbook(vlookup_file)
    vlookup_sheet_names = wvl.sheetnames

    #Iterate over worksheets in wb
    for sheet_name in wb.sheetnames:
        wb_sheet = wb[sheet_name]

        #Get the corresponding worksheet in wvl
        if sheet_name in vlookup_sheet_names:
            wvl_sheet = wvl[sheet_name]
        else:
            continue  #Skip if there is no matching worksheet in worksheet(wvl)

        #Iterate over rows in the current worksheet(wb)
        for row in wb_sheet.iter_rows(min_row=2):
            worker_id = row[7].value  #Value in column 8 (Worker ID)

            #Perform VLOOKUP in wvl
            for vlookup_row in wvl_sheet.iter_rows(min_row=2):
                if vlookup_row[7].value == worker_id:  #Match worker_id in column 8
                    comments = vlookup_row[26].value  #Value in column 27 (Comments)
                    row[26].value = comments  #Write comments to wb
                    row[26].font = Font(size=9, name='Arial Unicode MS')
                    break  # Break after finding the first match

    #Save workbook
    output_file = 'output.xlsx'
    writer.book.save(output_file)

    #Set filename and path for download button
    filename = st.text_input('Rename the file:', 'output.xlsx')
    st.success('Make sure to press "Enter" or click elsewhere after renaming the file', icon="✅")

    #Create download button with correct filename and path
    with open(output_file, 'rb') as f:
        data = f.read()
    b64 = base64.b64encode(data).decode('utf-8')
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{filename}">Download Excel file</a>'
    st.markdown(href, unsafe_allow_html=True)
